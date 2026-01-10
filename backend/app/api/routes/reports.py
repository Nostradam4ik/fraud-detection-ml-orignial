"""
Reports Routes - Generate and download PDF/Excel reports

Author: Zhmuryk Andrii
Copyright (c) 2024 - All Rights Reserved
"""

import io
from datetime import datetime, timedelta
from typing import Optional, Literal

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func

from ...db.database import get_db
from ...db.models import Prediction
from ...models.schemas import UserResponse
from ...services.auth_service import get_current_user

router = APIRouter(prefix="/reports", tags=["Reports"])


def generate_excel_report(
    predictions: list,
    user: UserResponse,
    title: str,
    include_summary: bool = True
) -> bytes:
    """Generate an Excel report with predictions data"""
    try:
        import pandas as pd
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils.dataframe import dataframe_to_rows
        from openpyxl.chart import PieChart, Reference, BarChart

        wb = Workbook()

        # Summary Sheet
        if include_summary and predictions:
            ws_summary = wb.active
            ws_summary.title = "Summary"

            # Title
            ws_summary['A1'] = title
            ws_summary['A1'].font = Font(size=18, bold=True, color="2563EB")
            ws_summary.merge_cells('A1:D1')

            ws_summary['A3'] = f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}"
            ws_summary['A4'] = f"User: {user.username}"
            ws_summary['A5'] = f"Total Predictions: {len(predictions)}"

            # Calculate stats
            total = len(predictions)
            fraud_count = sum(1 for p in predictions if p.get('is_fraud'))
            legitimate_count = total - fraud_count
            avg_amount = sum(p.get('amount', 0) for p in predictions) / total if total > 0 else 0

            # Risk distribution
            risk_low = sum(1 for p in predictions if p.get('risk_score', 0) < 25)
            risk_medium = sum(1 for p in predictions if 25 <= p.get('risk_score', 0) < 50)
            risk_high = sum(1 for p in predictions if 50 <= p.get('risk_score', 0) < 75)
            risk_critical = sum(1 for p in predictions if p.get('risk_score', 0) >= 75)

            # Summary table
            summary_header = ['Metric', 'Value']
            summary_data = [
                ['Total Predictions', total],
                ['Fraud Detected', fraud_count],
                ['Legitimate', legitimate_count],
                ['Fraud Rate', f"{fraud_count / total * 100:.2f}%" if total > 0 else "0%"],
                ['Average Amount', f"${avg_amount:.2f}"],
                ['', ''],
                ['Risk Distribution', ''],
                ['Low (0-24)', risk_low],
                ['Medium (25-49)', risk_medium],
                ['High (50-74)', risk_high],
                ['Critical (75-100)', risk_critical],
            ]

            # Header styling
            header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Write summary table
            for col, header in enumerate(summary_header, 1):
                cell = ws_summary.cell(row=7, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')

            for row_idx, row_data in enumerate(summary_data, 8):
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = thin_border
                    if row_data[0] == 'Risk Distribution' or row_data[0] == '':
                        cell.font = Font(bold=True)

            # Column widths
            ws_summary.column_dimensions['A'].width = 25
            ws_summary.column_dimensions['B'].width = 20

        # Data Sheet
        ws_data = wb.create_sheet("Predictions Data")

        # Headers
        headers = ['Date', 'Amount', 'Is Fraud', 'Fraud Probability', 'Risk Score', 'Confidence']
        header_fill = PatternFill(start_color="374151", end_color="374151", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        for col, header in enumerate(headers, 1):
            cell = ws_data.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        # Data rows
        fraud_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")

        for row_idx, pred in enumerate(predictions, 2):
            ws_data.cell(row=row_idx, column=1, value=pred.get('created_at', '')[:19] if pred.get('created_at') else '')
            ws_data.cell(row=row_idx, column=2, value=pred.get('amount', 0))
            ws_data.cell(row=row_idx, column=3, value='Yes' if pred.get('is_fraud') else 'No')
            ws_data.cell(row=row_idx, column=4, value=f"{pred.get('fraud_probability', 0) * 100:.2f}%")
            ws_data.cell(row=row_idx, column=5, value=pred.get('risk_score', 0))
            ws_data.cell(row=row_idx, column=6, value=pred.get('confidence', ''))

            # Highlight fraud rows
            if pred.get('is_fraud'):
                for col in range(1, 7):
                    ws_data.cell(row=row_idx, column=col).fill = fraud_fill

        # Column widths
        ws_data.column_dimensions['A'].width = 20
        ws_data.column_dimensions['B'].width = 12
        ws_data.column_dimensions['C'].width = 10
        ws_data.column_dimensions['D'].width = 18
        ws_data.column_dimensions['E'].width = 12
        ws_data.column_dimensions['F'].width = 12

        # Freeze header row
        ws_data.freeze_panes = 'A2'

        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    except ImportError as e:
        raise HTTPException(
            status_code=500,
            detail=f"Excel generation requires openpyxl. Error: {str(e)}"
        )


def generate_pdf_report(
    title: str,
    summary: dict,
    predictions: list,
    user: UserResponse,
    period_days: int
) -> bytes:
    """Generate a PDF report"""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.lib.enums import TA_CENTER, TA_LEFT

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        styles = getSampleStyleSheet()
        story = []

        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            spaceAfter=30,
            alignment=TA_CENTER
        )
        story.append(Paragraph(title, title_style))
        story.append(Spacer(1, 12))

        # Report info
        info_style = styles['Normal']
        story.append(Paragraph(f"<b>Generated:</b> {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}", info_style))
        story.append(Paragraph(f"<b>User:</b> {user.username}", info_style))
        story.append(Paragraph(f"<b>Period:</b> Last {period_days} days", info_style))
        story.append(Spacer(1, 20))

        # Summary section
        story.append(Paragraph("<b>Summary</b>", styles['Heading2']))
        story.append(Spacer(1, 10))

        summary_data = [
            ["Metric", "Value"],
            ["Total Predictions", str(summary.get("total_predictions", 0))],
            ["Fraud Detected", str(summary.get("fraud_count", 0))],
            ["Legitimate", str(summary.get("legitimate_count", 0))],
            ["Fraud Rate", f"{summary.get('fraud_rate', 0) * 100:.2f}%"],
            ["Avg Amount", f"${summary.get('avg_amount', 0):.2f}"],
            ["Trend", summary.get("trend", "stable").capitalize()],
        ]

        summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2563eb')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f8fafc')),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#e2e8f0')),
        ]))
        story.append(summary_table)
        story.append(Spacer(1, 20))

        # Risk Distribution
        story.append(Paragraph("<b>Risk Distribution</b>", styles['Heading2']))
        story.append(Spacer(1, 10))

        risk_dist = summary.get("risk_distribution", {})
        risk_data = [
            ["Risk Level", "Count"],
            ["Low (0-24)", str(risk_dist.get("low", 0))],
            ["Medium (25-49)", str(risk_dist.get("medium", 0))],
            ["High (50-74)", str(risk_dist.get("high", 0))],
            ["Critical (75-100)", str(risk_dist.get("critical", 0))],
        ]

        risk_table = Table(risk_data, colWidths=[3*inch, 2*inch])
        risk_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#dc2626')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#fef2f2')),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#fecaca')),
        ]))
        story.append(risk_table)
        story.append(Spacer(1, 20))

        # Recent Predictions (last 20)
        if predictions:
            story.append(Paragraph("<b>Recent Predictions</b>", styles['Heading2']))
            story.append(Spacer(1, 10))

            pred_data = [["Date", "Amount", "Fraud", "Probability", "Risk"]]
            for p in predictions[:20]:
                pred_data.append([
                    p["created_at"][:10] if p.get("created_at") else "",
                    f"${p.get('amount', 0):.2f}",
                    "Yes" if p.get("is_fraud") else "No",
                    f"{p.get('fraud_probability', 0) * 100:.1f}%",
                    str(p.get("risk_score", 0))
                ])

            pred_table = Table(pred_data, colWidths=[1.5*inch, 1*inch, 0.8*inch, 1*inch, 0.8*inch])
            pred_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#374151')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
                ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f9fafb')),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#d1d5db')),
            ]))
            story.append(pred_table)

        # Footer
        story.append(Spacer(1, 30))
        footer_style = ParagraphStyle(
            'Footer',
            parent=styles['Normal'],
            fontSize=8,
            textColor=colors.gray,
            alignment=TA_CENTER
        )
        story.append(Paragraph(
            "Generated by Fraud Detection ML System | © 2024 Zhmuryk Andrii",
            footer_style
        ))

        doc.build(story)
        buffer.seek(0)
        return buffer.getvalue()

    except ImportError:
        # ReportLab not installed, return simple text
        return b"PDF generation requires reportlab. Please install it with: pip install reportlab"


@router.get(
    "/fraud-summary",
    summary="Generate fraud summary report",
    description="Generate a PDF report with fraud detection summary."
)
async def generate_fraud_summary_report(
    days: int = Query(30, ge=1, le=365),
    current_user: UserResponse = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Generate a PDF fraud summary report"""
    end_date = datetime.utcnow()
    start_date = end_date - timedelta(days=days)

    # Get predictions
    predictions = (
        db.query(Prediction)
        .filter(
            Prediction.user_id == int(current_user.id),
            Prediction.created_at >= start_date
        )
        .order_by(Prediction.created_at.desc())
        .all()
    )

    # Calculate summary
    total = len(predictions)
    fraud = [p for p in predictions if p.is_fraud]
    legitimate = [p for p in predictions if not p.is_fraud]

    if total > 0:
        avg_amount = sum(p.amount for p in predictions) / total

        # Risk distribution
        risk_dist = {"low": 0, "medium": 0, "high": 0, "critical": 0}
        for p in predictions:
            if p.risk_score < 25:
                risk_dist["low"] += 1
            elif p.risk_score < 50:
                risk_dist["medium"] += 1
            elif p.risk_score < 75:
                risk_dist["high"] += 1
            else:
                risk_dist["critical"] += 1

        # Trend
        mid_date = start_date + timedelta(days=days // 2)
        first_half = [p for p in predictions if p.created_at < mid_date]
        second_half = [p for p in predictions if p.created_at >= mid_date]
        first_rate = len([p for p in first_half if p.is_fraud]) / len(first_half) if first_half else 0
        second_rate = len([p for p in second_half if p.is_fraud]) / len(second_half) if second_half else 0

        if second_rate > first_rate * 1.1:
            trend = "increasing"
        elif second_rate < first_rate * 0.9:
            trend = "decreasing"
        else:
            trend = "stable"

        summary = {
            "total_predictions": total,
            "fraud_count": len(fraud),
            "legitimate_count": len(legitimate),
            "fraud_rate": len(fraud) / total,
            "avg_amount": avg_amount,
            "risk_distribution": risk_dist,
            "trend": trend
        }
    else:
        summary = {
            "total_predictions": 0,
            "fraud_count": 0,
            "legitimate_count": 0,
            "fraud_rate": 0,
            "avg_amount": 0,
            "risk_distribution": {"low": 0, "medium": 0, "high": 0, "critical": 0},
            "trend": "stable"
        }

    # Convert predictions to dict
    pred_list = [
        {
            "created_at": p.created_at.isoformat() if p.created_at else None,
            "amount": p.amount,
            "is_fraud": p.is_fraud,
            "fraud_probability": p.fraud_probability,
            "risk_score": p.risk_score
        }
        for p in predictions
    ]

    # Generate PDF
    pdf_content = generate_pdf_report(
        title="Fraud Detection Report",
        summary=summary,
        predictions=pred_list,
        user=current_user,
        period_days=days
    )

    filename = f"fraud_report_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.pdf"

    return StreamingResponse(
        io.BytesIO(pdf_content),
        media_type="application/pdf",
        headers={
            "Content-Disposition": f"attachment; filename={filename}"
        }
    )


@router.get(
    "/trend-analysis",
    summary="Generate trend analysis report",
    description="Generate a PDF report with fraud trend analysis over time."
)
async def generate_trend_analysis_report(
    days: int = Query(30, ge=7, le=365),
    current_user: UserResponse = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Generate a PDF trend analysis report"""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.enums import TA_CENTER

    end_date = datetime.utcnow()
    start_date = end_date - timedelta(days=days)

    predictions = (
        db.query(Prediction)
        .filter(
            Prediction.user_id == int(current_user.id),
            Prediction.created_at >= start_date
        )
        .order_by(Prediction.created_at)
        .all()
    )

    # Group by day
    daily_stats = {}
    for p in predictions:
        day = p.created_at.strftime('%Y-%m-%d')
        if day not in daily_stats:
            daily_stats[day] = {'total': 0, 'fraud': 0, 'amount': 0}
        daily_stats[day]['total'] += 1
        daily_stats[day]['fraud'] += 1 if p.is_fraud else 0
        daily_stats[day]['amount'] += p.amount

    # Calculate weekly trends
    weekly_trends = []
    week_start = start_date
    while week_start < end_date:
        week_end = week_start + timedelta(days=7)
        week_preds = [p for p in predictions if week_start <= p.created_at < week_end]
        if week_preds:
            fraud_count = sum(1 for p in week_preds if p.is_fraud)
            weekly_trends.append({
                'week': week_start.strftime('%Y-%m-%d'),
                'total': len(week_preds),
                'fraud': fraud_count,
                'rate': fraud_count / len(week_preds) * 100
            })
        week_start = week_end

    # Generate PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    story = []

    title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=24, spaceAfter=30, alignment=TA_CENTER)
    story.append(Paragraph("Trend Analysis Report", title_style))
    story.append(Spacer(1, 12))

    story.append(Paragraph(f"<b>Generated:</b> {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}", styles['Normal']))
    story.append(Paragraph(f"<b>User:</b> {current_user.username}", styles['Normal']))
    story.append(Paragraph(f"<b>Period:</b> {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}", styles['Normal']))
    story.append(Spacer(1, 20))

    story.append(Paragraph("<b>Weekly Fraud Trends</b>", styles['Heading2']))
    story.append(Spacer(1, 10))

    if weekly_trends:
        trend_data = [["Week Starting", "Total", "Fraud", "Fraud Rate"]]
        for w in weekly_trends:
            trend_data.append([w['week'], str(w['total']), str(w['fraud']), f"{w['rate']:.1f}%"])

        trend_table = Table(trend_data, colWidths=[2*inch, 1.2*inch, 1.2*inch, 1.2*inch])
        trend_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#059669')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#ecfdf5')),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#a7f3d0')),
        ]))
        story.append(trend_table)
    else:
        story.append(Paragraph("No data available for the selected period.", styles['Normal']))

    story.append(Spacer(1, 20))
    story.append(Paragraph("<b>Daily Breakdown (Recent)</b>", styles['Heading2']))
    story.append(Spacer(1, 10))

    sorted_days = sorted(daily_stats.items(), reverse=True)[:14]
    if sorted_days:
        daily_data = [["Date", "Predictions", "Fraud", "Avg Amount"]]
        for day, stats in sorted_days:
            avg_amt = stats['amount'] / stats['total'] if stats['total'] > 0 else 0
            daily_data.append([day, str(stats['total']), str(stats['fraud']), f"${avg_amt:.2f}"])

        daily_table = Table(daily_data, colWidths=[1.8*inch, 1.2*inch, 1*inch, 1.2*inch])
        daily_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0284c7')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f0f9ff')),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#bae6fd')),
        ]))
        story.append(daily_table)

    story.append(Spacer(1, 20))
    story.append(Paragraph("<b>Trend Summary</b>", styles['Heading2']))
    story.append(Spacer(1, 10))

    if len(weekly_trends) >= 2:
        first_week_rate = weekly_trends[0]['rate']
        last_week_rate = weekly_trends[-1]['rate']
        if last_week_rate > first_week_rate * 1.2:
            trend_text = "INCREASING - Fraud rate has increased significantly."
            trend_color = colors.HexColor('#dc2626')
        elif last_week_rate < first_week_rate * 0.8:
            trend_text = "DECREASING - Fraud rate has decreased."
            trend_color = colors.HexColor('#059669')
        else:
            trend_text = "STABLE - Fraud rate remains relatively constant."
            trend_color = colors.HexColor('#0284c7')
        trend_style = ParagraphStyle('Trend', parent=styles['Normal'], textColor=trend_color, fontSize=12)
        story.append(Paragraph(trend_text, trend_style))
    else:
        story.append(Paragraph("Insufficient data for trend analysis.", styles['Normal']))

    story.append(Spacer(1, 30))
    footer_style = ParagraphStyle('Footer', parent=styles['Normal'], fontSize=8, textColor=colors.gray, alignment=TA_CENTER)
    story.append(Paragraph("Generated by Fraud Detection ML System | (c) 2024 Zhmuryk Andrii", footer_style))

    doc.build(story)
    buffer.seek(0)

    filename = f"trend_analysis_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.pdf"
    return StreamingResponse(buffer, media_type="application/pdf", headers={"Content-Disposition": f"attachment; filename={filename}"})


@router.get(
    "/high-risk",
    summary="Generate high risk transactions report",
    description="Generate a PDF report listing all high-risk transactions."
)
async def generate_high_risk_report(
    days: int = Query(30, ge=1, le=365),
    threshold: int = Query(50, ge=0, le=100),
    current_user: UserResponse = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Generate a PDF report of high-risk transactions"""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.enums import TA_CENTER

    end_date = datetime.utcnow()
    start_date = end_date - timedelta(days=days)

    predictions = (
        db.query(Prediction)
        .filter(
            Prediction.user_id == int(current_user.id),
            Prediction.created_at >= start_date,
            Prediction.risk_score >= threshold
        )
        .order_by(Prediction.risk_score.desc())
        .limit(100)
        .all()
    )

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    story = []

    title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=24, spaceAfter=30, alignment=TA_CENTER)
    story.append(Paragraph("High Risk Transactions Report", title_style))
    story.append(Spacer(1, 12))

    story.append(Paragraph(f"<b>Generated:</b> {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}", styles['Normal']))
    story.append(Paragraph(f"<b>User:</b> {current_user.username}", styles['Normal']))
    story.append(Paragraph(f"<b>Period:</b> Last {days} days", styles['Normal']))
    story.append(Paragraph(f"<b>Risk Threshold:</b> {threshold}+", styles['Normal']))
    story.append(Paragraph(f"<b>Total High-Risk Found:</b> {len(predictions)}", styles['Normal']))
    story.append(Spacer(1, 20))

    story.append(Paragraph("<b>Risk Level Distribution</b>", styles['Heading2']))
    story.append(Spacer(1, 10))

    critical = sum(1 for p in predictions if p.risk_score >= 75)
    high = sum(1 for p in predictions if 50 <= p.risk_score < 75)
    medium = sum(1 for p in predictions if 25 <= p.risk_score < 50)

    risk_data = [
        ["Risk Level", "Count", "Total Amount"],
        ["Critical (75-100)", str(critical), f"${sum(p.amount for p in predictions if p.risk_score >= 75):,.2f}"],
        ["High (50-74)", str(high), f"${sum(p.amount for p in predictions if 50 <= p.risk_score < 75):,.2f}"],
    ]
    if threshold < 50:
        risk_data.append(["Medium (25-49)", str(medium), f"${sum(p.amount for p in predictions if 25 <= p.risk_score < 50):,.2f}"])

    risk_table = Table(risk_data, colWidths=[2*inch, 1.5*inch, 2*inch])
    risk_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#dc2626')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#fef2f2')),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#fecaca')),
    ]))
    story.append(risk_table)
    story.append(Spacer(1, 20))

    story.append(Paragraph("<b>Transaction Details</b>", styles['Heading2']))
    story.append(Spacer(1, 10))

    if predictions:
        trans_data = [["Date", "Amount", "Risk", "Probability", "Fraud"]]
        for p in predictions[:50]:
            trans_data.append([
                p.created_at.strftime('%Y-%m-%d %H:%M') if p.created_at else "N/A",
                f"${p.amount:,.2f}",
                str(p.risk_score),
                f"{p.fraud_probability * 100:.1f}%",
                "Yes" if p.is_fraud else "No"
            ])

        trans_table = Table(trans_data, colWidths=[1.8*inch, 1.2*inch, 0.8*inch, 1*inch, 0.7*inch])
        trans_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#374151')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f9fafb')),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#d1d5db')),
        ]))
        story.append(trans_table)
        if len(predictions) > 50:
            story.append(Spacer(1, 10))
            story.append(Paragraph(f"<i>Showing 50 of {len(predictions)} transactions</i>", styles['Normal']))
    else:
        story.append(Paragraph("No high-risk transactions found.", styles['Normal']))

    story.append(Spacer(1, 30))
    footer_style = ParagraphStyle('Footer', parent=styles['Normal'], fontSize=8, textColor=colors.gray, alignment=TA_CENTER)
    story.append(Paragraph("Generated by Fraud Detection ML System | (c) 2024 Zhmuryk Andrii", footer_style))

    doc.build(story)
    buffer.seek(0)

    filename = f"high_risk_transactions_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.pdf"
    return StreamingResponse(buffer, media_type="application/pdf", headers={"Content-Disposition": f"attachment; filename={filename}"})


@router.get(
    "/model-performance",
    summary="Generate model performance report",
    description="Generate a PDF report with model performance metrics."
)
async def generate_model_performance_report(
    current_user: UserResponse = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Generate a PDF model performance report"""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.enums import TA_CENTER
    from ...models.ml_model import fraud_model

    model_info = fraud_model.model_info if fraud_model.is_loaded else {}

    all_predictions = db.query(Prediction).filter(Prediction.user_id == int(current_user.id)).all()

    total = len(all_predictions)
    fraud_predictions = sum(1 for p in all_predictions if p.is_fraud)
    high_confidence = sum(1 for p in all_predictions if p.confidence == 'high')
    medium_confidence = sum(1 for p in all_predictions if p.confidence == 'medium')
    low_confidence = sum(1 for p in all_predictions if p.confidence == 'low')
    avg_response_time = sum(p.prediction_time_ms for p in all_predictions) / total if total > 0 else 0

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    story = []

    title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=24, spaceAfter=30, alignment=TA_CENTER)
    story.append(Paragraph("Model Performance Report", title_style))
    story.append(Spacer(1, 12))

    story.append(Paragraph(f"<b>Generated:</b> {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}", styles['Normal']))
    story.append(Paragraph(f"<b>User:</b> {current_user.username}", styles['Normal']))
    story.append(Spacer(1, 20))

    story.append(Paragraph("<b>Model Information</b>", styles['Heading2']))
    story.append(Spacer(1, 10))

    model_data = [
        ["Property", "Value"],
        ["Model Type", model_info.get('model_type', 'Random Forest')],
        ["Model Version", model_info.get('version', '1.0.0')],
        ["Status", "Loaded" if fraud_model.is_loaded else "Not Loaded"],
        ["Features Used", "30 (Time, V1-V28, Amount)"],
    ]

    model_table = Table(model_data, colWidths=[2.5*inch, 3*inch])
    model_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#7c3aed')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f5f3ff')),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#ddd6fe')),
    ]))
    story.append(model_table)
    story.append(Spacer(1, 20))

    story.append(Paragraph("<b>Training Metrics</b>", styles['Heading2']))
    story.append(Spacer(1, 10))

    metrics_data = [
        ["Metric", "Value"],
        ["Accuracy", f"{model_info.get('accuracy', 0.9985) * 100:.2f}%"],
        ["Precision", f"{model_info.get('precision', 0.9456) * 100:.2f}%"],
        ["Recall", f"{model_info.get('recall', 0.7891) * 100:.2f}%"],
        ["F1 Score", f"{model_info.get('f1_score', 0.8605) * 100:.2f}%"],
        ["ROC AUC", f"{model_info.get('roc_auc', 0.9821) * 100:.2f}%"],
    ]

    metrics_table = Table(metrics_data, colWidths=[2.5*inch, 2*inch])
    metrics_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#059669')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#ecfdf5')),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#a7f3d0')),
    ]))
    story.append(metrics_table)
    story.append(Spacer(1, 20))

    story.append(Paragraph("<b>Usage Statistics</b>", styles['Heading2']))
    story.append(Spacer(1, 10))

    usage_data = [
        ["Statistic", "Value"],
        ["Total Predictions", str(total)],
        ["Fraud Detected", str(fraud_predictions)],
        ["Legitimate", str(total - fraud_predictions)],
        ["Detection Rate", f"{fraud_predictions / total * 100:.2f}%" if total > 0 else "N/A"],
        ["Avg Response Time", f"{avg_response_time:.2f} ms"],
    ]

    usage_table = Table(usage_data, colWidths=[2.5*inch, 2*inch])
    usage_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2563eb')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#eff6ff')),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#bfdbfe')),
    ]))
    story.append(usage_table)
    story.append(Spacer(1, 20))

    story.append(Paragraph("<b>Confidence Distribution</b>", styles['Heading2']))
    story.append(Spacer(1, 10))

    conf_data = [
        ["Confidence Level", "Count", "Percentage"],
        ["High", str(high_confidence), f"{high_confidence / total * 100:.1f}%" if total > 0 else "0%"],
        ["Medium", str(medium_confidence), f"{medium_confidence / total * 100:.1f}%" if total > 0 else "0%"],
        ["Low", str(low_confidence), f"{low_confidence / total * 100:.1f}%" if total > 0 else "0%"],
    ]

    conf_table = Table(conf_data, colWidths=[2*inch, 1.5*inch, 1.5*inch])
    conf_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#374151')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f9fafb')),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#d1d5db')),
    ]))
    story.append(conf_table)

    story.append(Spacer(1, 30))
    footer_style = ParagraphStyle('Footer', parent=styles['Normal'], fontSize=8, textColor=colors.gray, alignment=TA_CENTER)
    story.append(Paragraph("Generated by Fraud Detection ML System | (c) 2024 Zhmuryk Andrii", footer_style))

    doc.build(story)
    buffer.seek(0)

    filename = f"model_performance_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.pdf"
    return StreamingResponse(buffer, media_type="application/pdf", headers={"Content-Disposition": f"attachment; filename={filename}"})


@router.get(
    "/batch/{batch_id}",
    summary="Generate batch report",
    description="Generate a PDF report for a specific batch prediction."
)
async def generate_batch_report(
    batch_id: str,
    current_user: UserResponse = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Generate a PDF report for a batch prediction"""
    predictions = (
        db.query(Prediction)
        .filter(
            Prediction.user_id == int(current_user.id),
            Prediction.batch_id == batch_id
        )
        .order_by(Prediction.id)
        .all()
    )

    if not predictions:
        raise HTTPException(status_code=404, detail="Batch not found")

    total = len(predictions)
    fraud = [p for p in predictions if p.is_fraud]

    summary = {
        "total_predictions": total,
        "fraud_count": len(fraud),
        "legitimate_count": total - len(fraud),
        "fraud_rate": len(fraud) / total if total > 0 else 0,
        "avg_amount": sum(p.amount for p in predictions) / total if total > 0 else 0,
        "risk_distribution": {
            "low": sum(1 for p in predictions if p.risk_score < 25),
            "medium": sum(1 for p in predictions if 25 <= p.risk_score < 50),
            "high": sum(1 for p in predictions if 50 <= p.risk_score < 75),
            "critical": sum(1 for p in predictions if p.risk_score >= 75),
        },
        "trend": "batch"
    }

    pred_list = [
        {
            "created_at": p.created_at.isoformat() if p.created_at else None,
            "amount": p.amount,
            "is_fraud": p.is_fraud,
            "fraud_probability": p.fraud_probability,
            "risk_score": p.risk_score
        }
        for p in predictions
    ]

    pdf_content = generate_pdf_report(
        title=f"Batch Report - {batch_id[:8]}",
        summary=summary,
        predictions=pred_list,
        user=current_user,
        period_days=1
    )

    filename = f"batch_report_{batch_id[:8]}.pdf"

    return StreamingResponse(
        io.BytesIO(pdf_content),
        media_type="application/pdf",
        headers={
            "Content-Disposition": f"attachment; filename={filename}"
        }
    )


# Excel Export Endpoints

@router.get(
    "/export/excel",
    summary="Export predictions to Excel",
    description="Export all predictions to an Excel file with summary and data sheets."
)
async def export_predictions_excel(
    days: int = Query(30, ge=1, le=365),
    current_user: UserResponse = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Export predictions to Excel format"""
    end_date = datetime.utcnow()
    start_date = end_date - timedelta(days=days)

    predictions = (
        db.query(Prediction)
        .filter(
            Prediction.user_id == int(current_user.id),
            Prediction.created_at >= start_date
        )
        .order_by(Prediction.created_at.desc())
        .all()
    )

    pred_list = [
        {
            "created_at": p.created_at.isoformat() if p.created_at else None,
            "amount": p.amount,
            "is_fraud": p.is_fraud,
            "fraud_probability": p.fraud_probability,
            "risk_score": p.risk_score,
            "confidence": p.confidence
        }
        for p in predictions
    ]

    excel_content = generate_excel_report(
        predictions=pred_list,
        user=current_user,
        title=f"Fraud Detection Report - Last {days} Days"
    )

    filename = f"predictions_export_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return StreamingResponse(
        io.BytesIO(excel_content),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={filename}"
        }
    )


@router.get(
    "/export/excel/fraud-only",
    summary="Export fraud predictions to Excel",
    description="Export only fraudulent predictions to an Excel file."
)
async def export_fraud_predictions_excel(
    days: int = Query(30, ge=1, le=365),
    current_user: UserResponse = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Export only fraud predictions to Excel format"""
    end_date = datetime.utcnow()
    start_date = end_date - timedelta(days=days)

    predictions = (
        db.query(Prediction)
        .filter(
            Prediction.user_id == int(current_user.id),
            Prediction.created_at >= start_date,
            Prediction.is_fraud == True
        )
        .order_by(Prediction.created_at.desc())
        .all()
    )

    pred_list = [
        {
            "created_at": p.created_at.isoformat() if p.created_at else None,
            "amount": p.amount,
            "is_fraud": p.is_fraud,
            "fraud_probability": p.fraud_probability,
            "risk_score": p.risk_score,
            "confidence": p.confidence
        }
        for p in predictions
    ]

    excel_content = generate_excel_report(
        predictions=pred_list,
        user=current_user,
        title=f"Fraud Transactions - Last {days} Days"
    )

    filename = f"fraud_export_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return StreamingResponse(
        io.BytesIO(excel_content),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={filename}"
        }
    )


@router.get(
    "/export/excel/high-risk",
    summary="Export high-risk predictions to Excel",
    description="Export high-risk predictions (risk score >= threshold) to an Excel file."
)
async def export_high_risk_excel(
    days: int = Query(30, ge=1, le=365),
    threshold: int = Query(50, ge=0, le=100),
    current_user: UserResponse = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Export high-risk predictions to Excel format"""
    end_date = datetime.utcnow()
    start_date = end_date - timedelta(days=days)

    predictions = (
        db.query(Prediction)
        .filter(
            Prediction.user_id == int(current_user.id),
            Prediction.created_at >= start_date,
            Prediction.risk_score >= threshold
        )
        .order_by(Prediction.risk_score.desc())
        .all()
    )

    pred_list = [
        {
            "created_at": p.created_at.isoformat() if p.created_at else None,
            "amount": p.amount,
            "is_fraud": p.is_fraud,
            "fraud_probability": p.fraud_probability,
            "risk_score": p.risk_score,
            "confidence": p.confidence
        }
        for p in predictions
    ]

    excel_content = generate_excel_report(
        predictions=pred_list,
        user=current_user,
        title=f"High Risk Transactions (>={threshold}) - Last {days} Days"
    )

    filename = f"high_risk_export_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return StreamingResponse(
        io.BytesIO(excel_content),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={filename}"
        }
    )


@router.get(
    "/export/csv",
    summary="Export predictions to CSV",
    description="Export all predictions to a CSV file."
)
async def export_predictions_csv(
    days: int = Query(30, ge=1, le=365),
    current_user: UserResponse = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Export predictions to CSV format"""
    import pandas as pd

    end_date = datetime.utcnow()
    start_date = end_date - timedelta(days=days)

    predictions = (
        db.query(Prediction)
        .filter(
            Prediction.user_id == int(current_user.id),
            Prediction.created_at >= start_date
        )
        .order_by(Prediction.created_at.desc())
        .all()
    )

    data = [
        {
            "date": p.created_at.isoformat() if p.created_at else None,
            "amount": p.amount,
            "is_fraud": p.is_fraud,
            "fraud_probability": p.fraud_probability,
            "risk_score": p.risk_score,
            "confidence": p.confidence,
            "prediction_time_ms": p.prediction_time_ms
        }
        for p in predictions
    ]

    df = pd.DataFrame(data)
    output = io.StringIO()
    df.to_csv(output, index=False)
    output.seek(0)

    filename = f"predictions_export_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.csv"

    return StreamingResponse(
        io.BytesIO(output.getvalue().encode()),
        media_type="text/csv",
        headers={
            "Content-Disposition": f"attachment; filename={filename}"
        }
    )
